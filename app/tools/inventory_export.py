"""export_inventory_excel — export available units/inventory as an .xlsx sheet.

Triggered when the user asks to export available inventory / available units to Excel
(spreadsheet / xlsx). Queries individual project_units (one row per unit) with the same
filters as search_units, builds a formatted workbook, uploads it to S3 for a download
link, and — on Telegram — pushes the file straight into the chat. Mirrors the brochure
delivery path (S3 link + Telegram document; either alone is enough to succeed).
"""
import io
import logging
import uuid
from typing import Any, Optional

import boto3
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.brochures.storage import slugify
from app.config import settings
from app.db.models import Developer, Project, ProjectUnit
from app.tools.brochures import _send_telegram_document
from app.tools.registry import Tool, registry
from app.tools.units import _normalize_unit_types

log = logging.getLogger("askalpha.exports")

ASSETS_BUCKET = "assets-allegiance"
SQM_TO_SQFT = 10.7639
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MAX_ROWS = 5000  # keep the file snappy; tell the user if their query is wider than this

_s3 = boto3.client("s3", region_name=settings.s3_assets_region)

# (header, column width). One row per available unit.
_COLUMNS = [
    ("Project", 30), ("Developer", 22), ("City", 14), ("Community", 22),
    ("Type", 14), ("Beds", 7), ("Baths", 7), ("Size (sqft)", 12),
    ("Price (AED)", 15), ("Price/sqft (AED)", 16), ("View", 18),
    ("Floor", 8), ("Unit #", 12), ("Status", 12), ("Handover", 12),
]


def _f(v: Any) -> Optional[float]:
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


async def _query_units(db: AsyncSession, args: dict) -> tuple[list[dict], bool]:
    """Fetch individual available units matching the filters. Returns (rows, truncated)."""
    unit_types = _normalize_unit_types(args.get("unit_type"))
    price = func.coalesce(ProjectUnit.price, ProjectUnit.price_from)
    size = func.coalesce(ProjectUnit.size, ProjectUnit.size_from)
    area_unit = func.coalesce(
        func.nullif(func.lower(ProjectUnit.area_unit), "none"), Project.area_unit
    )

    stmt = (
        select(
            Project.name.label("project"),
            Developer.name.label("developer"),
            Project.city, Project.district, Project.completion_quarter,
            ProjectUnit.unit_type, ProjectUnit.bedrooms, ProjectUnit.bathrooms,
            size.label("size"), price.label("price"), ProjectUnit.price_per_area,
            ProjectUnit.view, ProjectUnit.floor, ProjectUnit.unit_number,
            ProjectUnit.status, area_unit.label("area_unit"),
        )
        .select_from(ProjectUnit)
        .join(Project, Project.id == ProjectUnit.project_id)
        .join(Developer, Developer.id == Project.developer_id, isouter=True)
        .where(Project.is_published == True)  # noqa: E712
    )

    pid = args.get("project_id")
    name = (args.get("project_name") or "").strip()
    if pid is not None:
        stmt = stmt.where(Project.id == int(pid))
    elif name:
        stmt = stmt.where(Project.name.ilike(f"%{name}%"))

    if unit_types:
        stmt = stmt.where(func.lower(ProjectUnit.unit_type).in_(unit_types))
    if args.get("bedrooms_min") is not None:
        stmt = stmt.where(ProjectUnit.bedrooms >= float(args["bedrooms_min"]))
    if args.get("bedrooms_max") is not None:
        stmt = stmt.where(ProjectUnit.bedrooms <= float(args["bedrooms_max"]))
    if args.get("min_size") is not None:
        stmt = stmt.where(size >= float(args["min_size"]))
    if args.get("max_size") is not None:
        stmt = stmt.where(size <= float(args["max_size"]))
    if args.get("min_unit_price") is not None:
        stmt = stmt.where(price >= float(args["min_unit_price"]))
    if args.get("max_unit_price") is not None:
        stmt = stmt.where(price <= float(args["max_unit_price"]))
    loc = (args.get("location") or "").strip()
    if loc:
        like = f"%{loc}%"
        stmt = stmt.where(
            Project.city.ilike(like) | Project.region.ilike(like)
            | Project.district.ilike(like) | Project.country.ilike(like)
        )

    stmt = stmt.order_by(Project.name, price.asc().nulls_last()).limit(MAX_ROWS + 1)
    rows = (await db.execute(stmt)).mappings().all()
    truncated = len(rows) > MAX_ROWS
    rows = rows[:MAX_ROWS]

    out: list[dict] = []
    for r in rows:
        size_v = _f(r["size"])
        if size_v and str(r["area_unit"] or "").startswith("sqm"):
            size_v *= SQM_TO_SQFT
        price_v = _f(r["price"])
        ppsf = _f(r["price_per_area"])
        if ppsf is None and price_v and size_v:
            ppsf = price_v / size_v
        beds = _f(r["bedrooms"])
        out.append({
            "project": r["project"], "developer": r["developer"], "city": r["city"],
            "community": r["district"], "type": (r["unit_type"] or "").title(),
            "beds": "Studio" if beds == 0 else (beds if beds is not None else None),
            "baths": _f(r["bathrooms"]),
            "size": round(size_v) if size_v else None,
            "price": round(price_v) if price_v else None,
            "ppsf": round(ppsf) if ppsf else None,
            "view": r["view"], "floor": _f(r["floor"]),
            "unit_number": r["unit_number"], "status": r["status"] or "available",
            "handover": r["completion_quarter"],
        })
    return out, truncated


def _build_xlsx(units: list[dict], title: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="8F7748")  # Allegiance accent
    for col, (label, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width

    for u in units:
        ws.append([
            u["project"], u["developer"], u["city"], u["community"], u["type"],
            u["beds"], u["baths"], u["size"], u["price"], u["ppsf"], u["view"],
            u["floor"], u["unit_number"], u["status"], u["handover"],
        ])

    # Number formats: AED with thousands separators, sizes as integers.
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=10):
        for cell in row:
            cell.number_format = "#,##0"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUMNS))}{ws.max_row}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _upload_xlsx(data: bytes, name: str) -> tuple[str, str]:
    """Store the xlsx and return (s3_key, presigned_url). Runs the blocking boto3 calls
    in a thread."""
    import asyncio

    key = f"generated/inventory/{slugify(name)}-{uuid.uuid4().hex[:8]}.xlsx"
    download = f"{slugify(name)}-inventory.xlsx"

    def _put_and_sign() -> str:
        _s3.put_object(
            Bucket=ASSETS_BUCKET, Key=key, Body=data, ContentType=XLSX_MIME,
            ContentDisposition=f'attachment; filename="{download}"',
        )
        return _s3.generate_presigned_url(
            "get_object", Params={"Bucket": ASSETS_BUCKET, "Key": key},
            ExpiresIn=7 * 24 * 3600,
        )

    url = await asyncio.to_thread(_put_and_sign)
    return key, url


def _scope_label(args: dict, units: list[dict]) -> str:
    """A human label for the file / caption from the active filters."""
    if args.get("project_name"):
        return str(args["project_name"])
    if units and len({u["project"] for u in units}) == 1:
        return units[0]["project"]
    bits = []
    if args.get("bedrooms_min") or args.get("bedrooms_max"):
        lo, hi = args.get("bedrooms_min"), args.get("bedrooms_max")
        bits.append(f"{int(lo)}BR" if lo == hi and lo is not None else "BR")
    types = _normalize_unit_types(args.get("unit_type"))
    if types:
        bits.append("/".join(types))
    if args.get("location"):
        bits.append(str(args["location"]))
    return " ".join(bits) or "Available inventory"


async def export_inventory_excel_handler(db: AsyncSession, args: dict, ctx: dict) -> dict:
    units, truncated = await _query_units(db, args)
    if not units:
        return {"status": "empty", "count": 0,
                "message": "No available units match that — broaden the filters or check the project name."}

    label = _scope_label(args, units)
    try:
        xlsx = _build_xlsx(units, label)
    except Exception as e:
        log.exception("xlsx build failed")
        return {"error": f"Couldn't build the Excel file: {e}"}

    filename = f"{slugify(label)}-inventory.xlsx"

    s3_key, url = None, None
    try:
        s3_key, url = await _upload_xlsx(xlsx, label)
    except Exception as e:
        log.error("inventory xlsx S3 upload failed (continuing with Telegram only): %s", e)

    delivered = False
    tg_chat_id = ctx.get("telegram_chat_id")
    if tg_chat_id:
        cap = f"📊 {label} — {len(units)} available unit{'s' if len(units) != 1 else ''}"
        delivered = await _send_telegram_document(
            int(tg_chat_id), xlsx, filename, cap, mime_type=XLSX_MIME,
        )

    if not delivered and not url:
        return {"error": "The Excel was built but couldn't be delivered: the S3 download link "
                         "needs an admin to grant s3:PutObject on the assets bucket, and no "
                         "Telegram chat is linked. Try again from Telegram, or ask an admin."}

    log.info("inventory export label=%r rows=%d telegram=%s url=%s", label, len(units), delivered, s3_key)
    result = {
        "status": "completed",
        "label": label,
        "row_count": len(units),
        "truncated": truncated,
        "xlsx_url": url,
        "filename": filename,
        "sent_to_telegram": delivered,
    }
    if truncated:
        result["note"] = f"Capped at {MAX_ROWS} rows — narrow the filters to capture the rest."
    elif url:
        result["url_expires"] = "7 days"
    return result


registry.register(Tool(
    name="export_inventory_excel",
    description=(
        "Export available units / inventory to an Excel (.xlsx) spreadsheet — one row per "
        "available unit (project, developer, location, type, beds, baths, size, price, "
        "price/sqft, view, floor, unit #, status, handover). Use this WHENEVER the user asks "
        "to export / download / 'give me an Excel/spreadsheet/sheet' of the available "
        "inventory or units, or to put search_units results into a file. Takes the SAME filters "
        "as search_units (unit_type, bedrooms_min/max, min/max_unit_price, min/max_size, "
        "location) and/or a project_name/project_id to export one project's full inventory. "
        "On Telegram the file is pushed straight into the chat; it also returns a download URL. "
        "Returns row_count and, if the result was capped, a truncated flag."
    ),
    input_schema={
        "type": "object",
        "properties": {
            "project_name": {"type": "string", "description": "Export one project's full inventory, by exact/contained name (e.g. 'Damac Islands')."},
            "project_id": {"type": "integer", "description": "Numeric project ID (alternative to project_name)."},
            "unit_type": {
                "type": "array",
                "items": {"type": "string", "enum": ["apartments", "villa", "townhouse", "duplex", "penthouse", "hotel apartments"]},
                "description": "Filter to these unit types (same mapping as search_units).",
            },
            "bedrooms_min": {"type": "integer", "description": "Minimum bedrooms (set min=max for an exact count)."},
            "bedrooms_max": {"type": "integer", "description": "Maximum bedrooms."},
            "min_unit_price": {"type": "number", "description": "Lower bound on unit price in AED."},
            "max_unit_price": {"type": "number", "description": "Upper bound on unit price in AED."},
            "min_size": {"type": "number", "description": "Minimum unit size in sqft."},
            "max_size": {"type": "number", "description": "Maximum unit size in sqft."},
            "location": {"type": "string", "description": "Location filter across city/region/district/country (e.g. 'Dubai Marina')."},
        },
        "required": [],
    },
    handler=export_inventory_excel_handler,
))
